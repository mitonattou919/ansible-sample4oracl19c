#         1         2         3         4         5         6         7
#1234567890123456789012345678901234567890123456789012345678901234567890123456789

# coding: utf-8

import openpyxl
import sys
import os

# Define Worksheet name.
ws_name_env     = 'env'
ws_name_parm    = 'parm'
ws_name_dir     = 'dir'
ws_name_pkg     = 'pkg'
ws_name_grp     = 'group'
ws_name_user    = 'user'
ws_name_kernel  = 'kernel'
ws_name_ulimit  = 'ulimit'
ws_name_service = 'service'

# Define file name.
inventory_file = 'hosts'
var_dir_host   = 'host_vars'
var_dir_group  = 'group_vars'

max_target = 16                # Max number of target host.
default_password = 'Seibi1234' # Default passwod of user.

# [ START write_file]
# Write strings to file.
#
# Args:
#   file_name: specify target file name.
#   text: text strings to write.
# Returns:
#   none
#
#         1         2         3         4         5         6         7
#1234567890123456789012345678901234567890123456789012345678901234567890123456789
#
def write_file(file_name, text):
    with open(file_name, mode='a') as f:
        f.write(text)

# [ END write_file]


# Get target Excel file name.
args = sys.argv
my_wb = openpyxl.load_workbook(args[1], data_only=True)

ws_env     = my_wb[ws_name_env]
ws_parm    = my_wb[ws_name_parm]
ws_pkg     = my_wb[ws_name_pkg]
ws_dir     = my_wb[ws_name_dir]
ws_grp     = my_wb[ws_name_grp]
ws_user    = my_wb[ws_name_user]
ws_kernel  = my_wb[ws_name_kernel]
ws_ulimit  = my_wb[ws_name_ulimit]
ws_service = my_wb[ws_name_service]

# Environment type.
## Get environment type.
env_type = ws_env.cell(2, 2).value

## Confirm exectuion environment is correct.
question_msg = ('You execute this script in ' + env_type 
                + ' environment.\nAre you sure? [y/N]: ')
user_answer = input(question_msg)

if user_answer != 'y':
    exit()



#         1         2         3         4         5         6         7
#1234567890123456789012345678901234567890123456789012345678901234567890123456789
#
# Get environment informations.
#
env_values = []

for col in ws_env.iter_cols(min_row=5, min_col=2, max_row=9):
    env_values.append([cell.value for cell in col])
    #if env_values[0][0] is None: break


#         1         2         3         4         5         6         7
#1234567890123456789012345678901234567890123456789012345678901234567890123456789
#
# Inventory file.
#
## File initialization.
if os.path.exists(inventory_file + '_' + env_type):
    os.remove(inventory_file + '_' + env_type)

## Create inventory file
for env_value in env_values:
    if env_value[0] != None:
        text = ('[' + str(env_value[1]) + ']\n' 
                + str(env_value[1]) 
                + ' ansible_host=' + str(env_value[2])
                + ' ansible_user=' + str(env_value[3])
                + ' ansible_password=' + str(env_value[4]) + '\n\n')
        write_file(inventory_file + '_' + env_type, text)


#         1         2         3         4         5         6         7
#1234567890123456789012345678901234567890123456789012345678901234567890123456789
#
# Variable files.
#
## Directory initialization.
if not os.path.exists(var_dir_host):
    os.makedirs('./' + var_dir_host)

#
## Get parm informations.
#
parm_values = []

for row in ws_parm.iter_rows(min_row=1, min_col=2):
    parm_values.append([cell.value for cell in row])

nw_if_flg = 'off'

for parm_value in parm_values:
    if parm_value[0] == 'parameter':
        for i in range(max_target):
            if parm_value[3 + i] is not None:
                if parm_value[3 + i] == env_values[i][0]:
                    if os.path.exists(var_dir_host + '/' + str(env_values[i][0]) + '.yml'):
                        os.remove(var_dir_host + '/' + str(env_values[i][0]) + '.yml')
                else:
                    print('Error: Target host is not match!')
                    exit()

    # Hostname
    elif parm_value[0] == 'hostname':
        for i in range(max_target):
            if i >= 3 and parm_value[i] is not None:
                text = ('\n# Hostname (roles/linux_1st/tasks/hostname.yml)\n'
                        + str(parm_value[1]) + ' \'' + str(parm_value[i]) + '\'\n') 
                write_file(var_dir_host + '/' + str(env_values[i - 3][0]) + '.yml', text)

    # SELinux
    elif parm_value[0] == 'selinux':
        for i in range(max_target):
            if i >= 3 and parm_value[i] is not None:
                text = ('\n# SELinux (roles/linux_1st/tasks/selinux.yml)\n'
                        + str(parm_value[1]) + ' \'' + str(parm_value[i]) + '\'\n') 
                write_file(var_dir_host + '/' + str(env_values[i - 3][0]) + '.yml', text)

    # Timezone
    elif parm_value[0] == 'timezone':
        for i in range(max_target):
            if i >= 3 and parm_value[i] is not None:
                text = ('\n# Timezone (roles/linux_1st/tasks/timezone.yml)\n'
                        + str(parm_value[1]) + ' \'' + str(parm_value[i]) + '\'\n') 
                write_file(var_dir_host + '/' + str(env_values[i - 3][0]) + '.yml', text)

    # Network Interface
    elif parm_value[0] == 'nw_if':
        for i in range(max_target):
            if i >= 3 and parm_value[i] is not None and str(parm_value[i]) == '-':
                text = ('\n# Network Interface (roles/linux_1st/tasks/nmcli.yml)\n'
                        + str(parm_value[1]) + '\n')
                write_file(var_dir_host + '/' + str(env_values[i - 3][0]) + '.yml', text)
            elif i >= 3 and parm_value[i] is not None:
                text = (' - { ' + str(parm_value[i]) + ' }\n')
                write_file(var_dir_host + '/' + str(env_values[i - 3][0]) + '.yml', text)

    # DNS Client
    elif parm_value[0] == 'dns_cli':
        for i in range(max_target):
            if i >= 3 and parm_value[i] is not None:
                text = ('\n# DNS Client (roles/linux_1st/tasks/dns_client.yml)\n'
                        + str(parm_value[1]) + '\n - { ' + str(parm_value[i]) + ' }\n') 
                write_file(var_dir_host + '/' + str(env_values[i - 3][0]) + '.yml', text)

    # Partition
    elif parm_value[0] == 'part':
        for i in range(max_target):
            if i >= 3 and parm_value[i] is not None and str(parm_value[i]) == '-':
                text = ('\n# Partition (roles/linux_1st/tasks/parted.yml)\n'
                        + str(parm_value[1]) + '\n')
                write_file(var_dir_host + '/' + str(env_values[i - 3][0]) + '.yml', text)
            elif i >= 3 and parm_value[i] is not None:
                text = (' - { ' + str(parm_value[i]) + ' }\n')
                write_file(var_dir_host + '/' + str(env_values[i - 3][0]) + '.yml', text)

    # LVM2 - VG
    elif parm_value[0] == 'vg':
        for i in range(max_target):
            if i >= 3 and parm_value[i] is not None and str(parm_value[i]) == '-':
                text = ('\n# LVM2 - VG (roles/linux_1st/tasks/vg.yml)\n'
                        + str(parm_value[1]) + '\n')
                write_file(var_dir_host + '/' + str(env_values[i - 3][0]) + '.yml', text)
            elif i >= 3 and parm_value[i] is not None:
                text = (' - { ' + str(parm_value[i]) + ' }\n')
                write_file(var_dir_host + '/' + str(env_values[i - 3][0]) + '.yml', text)

    # LVM2 - LV
    elif parm_value[0] == 'lv':
        for i in range(max_target):
            if i >= 3 and parm_value[i] is not None and str(parm_value[i]) == '-':
                text = ('\n# LVM2 - LV (roles/linux_1st/tasks/lvol.yml)\n'
                        + str(parm_value[1]) + '\n')
                write_file(var_dir_host + '/' + str(env_values[i - 3][0]) + '.yml', text)
            elif i >= 3 and parm_value[i] is not None:
                text = (' - { ' + str(parm_value[i]) + ' }\n')
                write_file(var_dir_host + '/' + str(env_values[i - 3][0]) + '.yml', text)

    # Filesystem Formatting
    elif parm_value[0] == 'mkfs':
        for i in range(max_target):
            if i >= 3 and parm_value[i] is not None and str(parm_value[i]) == '-':
                text = ('\n# Filesystem Formatting (roles/linux_1st/tasks/filesystem.yml)\n'
                        + str(parm_value[1]) + '\n')
                write_file(var_dir_host + '/' + str(env_values[i - 3][0]) + '.yml', text)
            elif i >= 3 and parm_value[i] is not None:
                text = (' - { ' + str(parm_value[i]) + ' }\n')
                write_file(var_dir_host + '/' + str(env_values[i - 3][0]) + '.yml', text)

    # fstab
    elif parm_value[0] == 'fstab':
        for i in range(max_target):
            if i >= 3 and parm_value[i] is not None and str(parm_value[i]) == '-':
                text = ('\n# fstab configuration (roles/linux_1st/tasks/fstab.yml)\n'
                        + str(parm_value[1]) + '\n')
                write_file(var_dir_host + '/' + str(env_values[i - 3][0]) + '.yml', text)
            elif i >= 3 and parm_value[i] is not None:
                text = (' - { ' + str(parm_value[i]) + ' }\n')
                write_file(var_dir_host + '/' + str(env_values[i - 3][0]) + '.yml', text)

    # mount-point
    elif parm_value[0] == 'dir_mp':
        for i in range(max_target):
            if i >= 3 and parm_value[i] is not None and str(parm_value[i]) == '-':
                text = ('\n# Mount-point directory configuration (roles/linux_1st/tasks/dir_mountpoint.yml)\n'
                        + str(parm_value[1]) + '\n')
                write_file(var_dir_host + '/' + str(env_values[i - 3][0]) + '.yml', text)
            elif i >= 3 and parm_value[i] is not None:
                text = (' - { ' + str(parm_value[i]) + ' }\n')
                write_file(var_dir_host + '/' + str(env_values[i - 3][0]) + '.yml', text)


#         1         2         3         4         5         6         7
#1234567890123456789012345678901234567890123456789012345678901234567890123456789
#
# Package
#
## Get package informations.
#
values = []

for row in ws_pkg.iter_rows(min_row=1, min_col=1):
    values.append([cell.value for cell in row])


for value in values:
    if value[0] == 'package':
        for i in range(max_target):
            if i >= 4 and value[i] is not None:
                if value[i] == env_values[i - 4][0]:
                    text = ('\n# Package installations (roles/linux_1st/tasks/dnf.yml)\n'
                            + 'linux_packages:\n')
                    write_file(var_dir_host + '/' + str(env_values[i - 4][0]) + '.yml', text)
                else:
                    print('Error: Target host is not match!')
                    exit()
    else:
        for i in range(max_target):
            if i >= 4 and value[i] is not None and value[i] == 'yes':
                text = (' - ' + str(value[0]) + '.' + str(value[1]) + '\n') 
                write_file(var_dir_host + '/' + str(env_values[i - 4][0]) + '.yml', text)


#         1         2         3         4         5         6         7
#1234567890123456789012345678901234567890123456789012345678901234567890123456789
#
# Directory
#
## Get directory informations.
#
values = []

for row in ws_dir.iter_rows(min_row=1, min_col=1):
    values.append([cell.value for cell in row])

for value in values:
    if value[0] == 'path':
        for i in range(max_target):
            if i >= 4 and value[i] is not None:
                if value[i] == env_values[i - 4][0]:
                    text = ('\n# Directory configurations (roles/linux_2nd/tasks/dir.yml)\n'
                            + 'linux_dir_info:\n')
                    write_file(var_dir_host + '/' + str(env_values[i - 4][0]) + '.yml', text)
                else:
                    print('Error: Target host is not match!')
                    exit()
    else:
        for i in range(max_target):
            if i >= 4 and value[i] is not None and value[i] == 'yes':
                text = (' - { dir_path: \'' + str(value[0]) + '\', '
                        + 'dir_owner: \'' + str(value[1]) + '\', ' 
                        + 'dir_group: \'' + str(value[2]) + '\', '
                        + 'dir_mode: \'' + str(value[3]) + '\' }\n')
                write_file(var_dir_host + '/' + str(env_values[i - 4][0]) + '.yml', text)

#         1         2         3         4         5         6         7
#1234567890123456789012345678901234567890123456789012345678901234567890123456789
#
# Group
#
values = []

for row in ws_grp.iter_rows(min_row=1, min_col=1):
    values.append([cell.value for cell in row])

for value in values:
    if value[0] == 'group':
        for i in range(max_target):
            if i >= 4 and value[i] is not None:
                if value[i] == env_values[i - 4][0]:
                    text = ('\n# Group configurations (roles/linux_2nd/tasks/group.yml)\n'
                            + 'linux_group_info:\n')
                    write_file(var_dir_host + '/' + str(env_values[i - 4][0]) + '.yml', text)
                else:
                    print('Error: Target host is not match!')
                    exit()
    else:
        for i in range(max_target):
            if i >= 4 and value[i] is not None and value[i] == 'yes':
                text = (' - { group_name: \'' + str(value[0]) + '\', '
                        + 'group_id: \'' + str(value[1]) + '\' }\n')
                write_file(var_dir_host + '/' + str(env_values[i - 4][0]) + '.yml', text)

#         1         2         3         4         5         6         7
#1234567890123456789012345678901234567890123456789012345678901234567890123456789
#
# User
#
values = []

for row in ws_user.iter_rows(min_row=1, min_col=1):
    values.append([cell.value for cell in row])

for value in values:
    if value[0] == 'user':
        for i in range(max_target):
            if i >= 4 and value[i] is not None:
                if value[i] == env_values[i - 4][0]:
                    text = ('\n# User configurations (roles/linux_2nd/tasks/user.yml)\n'
                            + 'linux_user_info:\n')
                    write_file(var_dir_host + '/' + str(env_values[i - 4][0]) + '.yml', text)
                else:
                    print('Error: Target host is not match!')
                    exit()
    else:
        for i in range(max_target):
            if i >= 4 and value[i] is not None and value[i] == 'yes':
                text = (' - { user_name: \'' + str(value[0]) + '\', '
                        + 'user_id: \'' + str(value[1]) + '\', ' 
                        + 'user_group: \'' + str(value[2]) + '\', \n'
                        + '     user_groups: \'' + str(value[3]) + '\', \n'
                        + '     user_password: \"{{ \'' + default_password 
                        + '\' | password_hash(\'sha512\', \'salt\') }}\"'
                        + ' }\n')
                write_file(var_dir_host + '/' + str(env_values[i - 4][0]) + '.yml', text)

#         1         2         3         4         5         6         7
#1234567890123456789012345678901234567890123456789012345678901234567890123456789
#
# Kernel
#
values = []

for row in ws_kernel.iter_rows(min_row=1, min_col=1):
    values.append([cell.value for cell in row])

for value in values:
    if value[0] == 'parameter':
        for i in range(max_target):
            if i >= 4 and value[i] is not None:
                if value[i] == env_values[i - 4][0]:
                    text = ('\n# Kernel Tunings (roles/linux_2nd/tasks/sysctl.yml)\n'
                            + 'linux_sysctl_info:\n')
                    write_file(var_dir_host + '/' + str(env_values[i - 4][0]) + '.yml', text)
                else:
                    print('Error: Target host is not match!')
                    exit()
    else:
        for i in range(max_target):
            if i >= 4 and value[i] is not None:
                text = (' - { sysctl_name: \'' + str(value[0]) + '\', '
                        + 'sysctl_value: \'' + str(value[i]) + '\' }\n')
                write_file(var_dir_host + '/' + str(env_values[i - 4][0]) + '.yml', text)

#         1         2         3         4         5         6         7
#1234567890123456789012345678901234567890123456789012345678901234567890123456789
#
# User limit
#
values = []

for row in ws_ulimit.iter_rows(min_row=1, min_col=1):
    values.append([cell.value for cell in row])

for value in values:
    if value[0] == 'domain':
        for i in range(max_target):
            if i >= 4 and value[i] is not None:
                if value[i] == env_values[i - 4][0]:
                    text = ('\n# User limit Tunings (roles/linux_2nd/tasks/ulimit.yml)\n'
                            + 'linux_ulimit_info:\n')
                    write_file(var_dir_host + '/' + str(env_values[i - 4][0]) + '.yml', text)
                else:
                    print('Error: Target host is not match!')
                    exit()
    else:
        for i in range(max_target):
            if i >= 4 and value[i] is not None:
                text = (' - { domain: \'' + str(value[0]) + '\', '
                        + 'limit_item: \'' + str(value[1]) + '\', '
                        + 'limit_type: \'' + str(value[2]) + '\', '
                        + 'value: \'' + str(value[i]) + '\' }\n')
                write_file(var_dir_host + '/' + str(env_values[i - 4][0]) + '.yml', text)


#         1         2         3         4         5         6         7
#1234567890123456789012345678901234567890123456789012345678901234567890123456789
#
# Service
#
values = []

for row in ws_service.iter_rows(min_row=1, min_col=1):
    values.append([cell.value for cell in row])

for value in values:
    if value[0] == 'service':
        for i in range(max_target):
            if i >= 4 and value[i] is not None:
                if value[i] == env_values[i - 4][0]:
                    text = ('\n# Service configurations (roles/linux_2nd/tasks/systemd.yml)\n'
                            + 'linux_systemd_info:\n')
                    write_file(var_dir_host + '/' + str(env_values[i - 4][0]) + '.yml', text)
                else:
                    print('Error: Target host is not match!')
                    exit()
    else:
        for i in range(max_target):
            if i >= 4 and value[i] is not None and value[i] == 'yes':
                text = (' - { service_name: \'' + str(value[0]) + '\', '
                        + 'service_enabled: \'' + str(value[1]) + '\' }\n')
                write_file(var_dir_host + '/' + str(env_values[i - 4][0]) + '.yml', text)
